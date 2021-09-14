import openpyxl as xl
import config as cfg

def findColumnNumber(ws, sval):
    for i in range(1, 30):
        if sval in ws.cell(1, i).value:
            return i

def getTotalRows(ws):
    for i in range(1, 1000):
        if ws.cell(i, 1).value == None:
            return i - 1

def loadExcels(filename):
    load_wb = xl.load_workbook(filename, data_only=True)
    load_ws = load_wb['설문지 응답 시트1']

    # 이름 열 번호
    col_name = findColumnNumber(load_ws, cfg.STRING_NAME)
    # 학번 열 번호
    col_id = findColumnNumber(load_ws, cfg.STRING_ID)
    # 시간대 열 번호
    col_time = findColumnNumber(load_ws, cfg.STRING_TIME)

    totalRows = getTotalRows(load_ws)

    with open('list.txt', 'w', encoding='UTF-8') as f:
        enrollList = []
        for i in range(2, totalRows + 1):
            row_name = load_ws.cell(i, col_name).value
            row_id = load_ws.cell(i, col_id).value
            row_mon = load_ws.cell(i, col_time).value
            row_tue = load_ws.cell(i, col_time+1).value
            row_wed = load_ws.cell(i, col_time+2).value
            row_thu = load_ws.cell(i, col_time+3).value
            row_fri = load_ws.cell(i, col_time+4).value
            row_sat = load_ws.cell(i, col_time+5).value
            row_sun = load_ws.cell(i, col_time+6).value

            timesstr = ''
            if row_mon != None:
                spls = row_mon.split(', ')
                for s in spls:
                    if s[0] in list(map(str, range(1, cfg.TIME_LIMIT + 1))):
                        timesstr += '월'+s[0]+' '
            if row_tue != None:
                spls = row_tue.split(', ')
                for s in spls:
                    if s[0] in list(map(str, range(1, cfg.TIME_LIMIT + 1))):
                        timesstr += '화'+s[0]+' '
            if row_wed != None:
                spls = row_wed.split(', ')
                for s in spls:
                    if s[0] in list(map(str, range(1, cfg.TIME_LIMIT + 1))):
                        timesstr += '수'+s[0]+' '
            if row_thu != None:
                spls = row_thu.split(', ')
                for s in spls:
                    if s[0] in list(map(str, range(1, cfg.TIME_LIMIT + 1))):
                        timesstr += '목'+s[0]+' '
            if row_fri != None:
                spls = row_fri.split(', ')
                for s in spls:
                    if s[0] in list(map(str, range(1, cfg.TIME_LIMIT + 1))):
                        timesstr += '금'+s[0]+' '
            if row_sat != None:
                spls = row_sat.split(', ')
                for s in spls:
                    if s[0] in list(map(str, range(1, cfg.TIME_LIMIT + 1))):
                        timesstr += '토'+s[0]+' '
            if row_sun != None:
                spls = row_sun.split(', ')
                for s in spls:
                    if s[0] in list(map(str, range(1, cfg.TIME_LIMIT + 1))):
                        timesstr += '일'+s[0]+' '

            enrollList.append(f'{row_name} {row_id} {timesstr}\n')
        f.writelines(enrollList)